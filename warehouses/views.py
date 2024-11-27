import os
from django.conf import settings
from django.views.generic import ListView, CreateView, DeleteView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import Warehouse, Coming, ProductArrival, ProductStock, WarehouseTransfer
from directory.models import Product
from django.urls import reverse_lazy
from .forms import ComingForm, WarehouseTransferForm
from django.urls import reverse
from django.views import View
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import mm
from reportlab.graphics.barcode import code39
from reportlab.lib.units import mm
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from django.db.models import Sum, F, Q
import requests
import openpyxl
from openpyxl.styles import Font
from django.shortcuts import redirect
from django.contrib import messages
from django.db.models import ExpressionWrapper, DecimalField
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.template.loader import render_to_string
from weasyprint import HTML


class ComingListView(LoginRequiredMixin, ListView):
    model = Coming
    template_name = 'warehouses/coming_list.html'
    context_object_name = 'coming_list'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['warehouses'] = Warehouse.objects.all()
        return context


class ComingCreateView(CreateView):
    model = Coming
    form_class = ComingForm
    template_name = 'warehouses/coming_list.html'

    def form_valid(self, form):
        self.object = form.save()
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('warehouses:coming-detail', kwargs={'pk': self.object.pk})


class ComingDetailView(LoginRequiredMixin, DetailView):
    model = Coming
    template_name = 'warehouses/coming_detail.html'
    context_object_name = 'coming'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        coming = self.get_object()
        vat_multiplier = 1 + (coming.vat_percentage / 100)

        product_arrivals = ProductArrival.objects.filter(coming=coming).annotate(
            unit_price_with_vat=ExpressionWrapper(
                F('unit_price') * vat_multiplier,
                output_field=DecimalField(max_digits=10, decimal_places=2)
            ),
            total_price_with_vat=ExpressionWrapper(
                F('quantity') * F('unit_price') * vat_multiplier,
                output_field=DecimalField(max_digits=15, decimal_places=2)
            )
        )

        total_quantity = product_arrivals.aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        total_amount = product_arrivals.aggregate(total_amount=Sum(F('quantity') * F('unit_price')))[
                           'total_amount'] or 0
        total_amount_with_vat = product_arrivals.aggregate(
            total_amount_with_vat=Sum('total_price_with_vat')
        )['total_amount_with_vat'] or 0

        context['products'] = Product.objects.all()
        context['product_arrivals'] = product_arrivals
        context['total_quantity'] = total_quantity
        context['total_amount'] = total_amount
        context['total_amount_with_vat'] = total_amount_with_vat

        return context

    @method_decorator(csrf_exempt)
    def post(self, request, *args, **kwargs):
        coming = self.get_object()
        if not coming.is_posted:
            coming.is_posted = True
            coming.save()
            return JsonResponse(
                {'success': True, 'message': 'Проводка успешно выполнена. Добавление новых товаров теперь недоступно.'})
        return JsonResponse({'success': False, 'message': 'Этот приход уже проведен.'})


class ComingDeleteView(DeleteView):
    model = Coming
    success_url = reverse_lazy('warehouses:coming-list')
    template_name = 'warehouses/coming_list.html'


class ProductArrivalCreateView(CreateView):
    model = ProductArrival
    fields = ['product', 'quantity', 'unit_price']
    template_name = 'warehouses/coming_detail.html'

    def get_success_url(self):
        coming_id = self.kwargs.get('coming_id')
        return reverse('warehouses:coming-detail', kwargs={'pk': coming_id})

    def form_valid(self, form):
        coming_id = self.kwargs.get('coming_id')
        coming = Coming.objects.get(id=coming_id)
        form.instance.coming = coming
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        coming_id = self.kwargs.get('coming_id')
        coming = Coming.objects.get(id=coming_id)
        context['coming'] = coming
        context['product_arrivals'] = ProductArrival.objects.filter(coming=coming)
        return context


class ProductArrivalDeleteView(DeleteView):
    model = ProductArrival
    template_name = 'warehouses/coming_detail.html'

    def get_success_url(self):
        product_arrival = self.get_object()
        coming = product_arrival.coming
        return reverse('warehouses:coming-detail', kwargs={'pk': coming.id})


class ProductArrivalDownloadPDFView(View):
    def get(self, request, *args, **kwargs):
        product_arrival = ProductArrival.objects.get(id=kwargs['pk'])
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="product_arrival_{product_arrival.id}.pdf"'
        width, height = 80 * mm, 40 * mm
        p = canvas.Canvas(response, pagesize=(width, height))
        font_path = os.path.join(settings.BASE_DIR, 'warehouses/static/fonts/DejaVuSans.ttf')
        pdfmetrics.registerFont(TTFont('DejaVu', font_path))
        product_name = f"{product_arrival.product.name}"
        p.setFont("DejaVu", 7)
        product_name_x = 13 * mm
        product_name_y = height - 15 * mm
        p.drawString(product_name_x, product_name_y, product_name)
        logo_path = os.path.join(settings.BASE_DIR, 'warehouses/static/warehouses/images/logo.png')
        logo_width = 15 * mm
        logo_height = 15 * mm
        logo_x = width - logo_width - 1 * mm
        logo_y = height - logo_height - -3.5 * mm

        if os.path.exists(logo_path):
            p.drawImage(logo_path, logo_x, logo_y, width=logo_width, height=logo_height, preserveAspectRatio=True,
                        mask='auto')
        else:
            print(f"Логотип не найден по пути: {logo_path}")

        barcode = code39.Extended39(product_arrival.barcode, barWidth=0.2 * mm, barHeight=10 * mm)
        barcode_width = barcode.width
        barcode_height = barcode.height
        barcode_x = (width - barcode_width) / 2
        barcode_y = 7.5 * mm
        barcode.drawOn(p, barcode_x, barcode_y)
        text_width = p.stringWidth(product_arrival.barcode, "Helvetica", 10)
        text_x = (width - text_width) / 2
        text_y = barcode_y - 5 * mm
        p.setFont("Helvetica", 10)
        p.drawString(text_x, text_y, product_arrival.barcode)
        p.showPage()
        p.save()

        return response


class ProductStockListView(LoginRequiredMixin, ListView):
    model = ProductStock
    template_name = "warehouses/leftovers_list.html"
    context_object_name = "stocks"

    def get_queryset(self):
        queryset = ProductStock.objects.select_related('warehouse', 'product')
        warehouse_id = self.request.GET.get('warehouse')
        product_query = self.request.GET.get('product')

        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)
        if product_query:
            queryset = queryset.filter(
                Q(product__name__icontains=product_query) | Q(barcode__icontains=product_query)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['warehouses'] = Warehouse.objects.all()
        queryset = self.get_queryset()
        total_quantity = queryset.aggregate(Sum('quantity'))['quantity__sum'] or 0
        context['total_quantity'] = total_quantity

        return context


class ProductStockExportView(View):
    TELEGRAM_BOT_TOKEN = "7867735433:AAEOO4CII3sBDwBfKUWy4FapOsNLgKRl0Vc"
    TELEGRAM_CHAT_ID = "-4542685446"

    def get(self, request, *args, **kwargs):
        warehouse_id = request.GET.get('warehouse')
        product_query = request.GET.get('product')

        queryset = ProductStock.objects.select_related('warehouse', 'product')
        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)
        if product_query:
            queryset = queryset.filter(
                Q(product__name__icontains=product_query) | Q(barcode__icontains=product_query)
            )

        total_quantity = queryset.aggregate(Sum('quantity'))['quantity__sum'] or 0
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Остатки'
        sheet.merge_cells('A1:D1')
        sheet['A1'] = 'Остатки по складам'
        sheet['A1'].font = Font(size=14, bold=True)
        sheet.append([])
        sheet.append(['Склад', 'Филиал', 'Товар', 'Штрих-код', 'Количество'])

        for stock in queryset:
            sheet.append([
                stock.warehouse.name,
                stock.warehouse.branch.name if stock.warehouse.branch else 'N/A',
                stock.product.name,
                stock.barcode,
                stock.quantity,
            ])

        sheet.append([])
        sheet.append(['', '', '', 'Итого', total_quantity])
        sheet.cell(sheet.max_row, 4).font = Font(bold=True)

        file_path = '/tmp/Остатки.xlsx'
        workbook.save(file_path)

        with open(file_path, 'rb') as f:
            url = f'https://api.telegram.org/bot{self.TELEGRAM_BOT_TOKEN}/sendDocument'
            response_telegram = requests.post(url, data={'chat_id': self.TELEGRAM_CHAT_ID}, files={'document': f})

        if response_telegram.status_code != 200:
            messages.error(request, f"Ошибка отправки файла в Telegram: {response_telegram.text}")
            return redirect('warehouses:product_stock_list')

        messages.success(request, "Файл успешно отправлен в Telegram")
        return redirect('warehouses:product_stock_list')


class WarehouseTransferListView(LoginRequiredMixin, ListView):
    model = WarehouseTransfer
    template_name = 'warehouses/transfer_list.html'
    context_object_name = 'transfers'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['warehouses'] = Warehouse.objects.all()
        context['products'] = Product.objects.all()
        return context


class WarehouseTransferCreateView(LoginRequiredMixin, CreateView):
    model = WarehouseTransfer
    form_class = WarehouseTransferForm
    template_name = 'warehouses/transfer_list.html'
    success_url = reverse_lazy('warehouses:transfer_list')

    def form_valid(self, form):
        source_warehouse = form.cleaned_data['source_warehouse']
        product = form.cleaned_data['product']
        quantity = form.cleaned_data['quantity']

        try:
            product_stock = ProductStock.objects.get(warehouse=source_warehouse, product=product)
        except ProductStock.DoesNotExist:
            messages.error(self.request, f"На складе '{source_warehouse.name}' отсутствует товар '{product.name}'.")
            return redirect('warehouses:transfer_list')

        if product_stock.quantity < quantity:
            messages.error(
                self.request,
                f"Недостаточно товара '{product.name}' на складе '{source_warehouse.name}'. "
                f"Доступно: {product_stock.quantity}."
            )
            return redirect('warehouses:transfer_list')

        return super().form_valid(form)


def generate_pdf(request, coming_id):
    try:
        coming = Coming.objects.get(id=coming_id)
    except Coming.DoesNotExist:
        return HttpResponse("Приход не найден", status=404)

    # Расчёты
    products = coming.products.all()
    total_amount = sum(product.quantity * product.unit_price for product in products)  # Общее количество денег
    vat_amount = total_amount * (coming.vat_percentage / 100)  # Сумма НДС

    # Отладка
    print(f"Coming ID: {coming.id}")
    print(f"Total Amount: {total_amount}")
    print(f"VAT Amount: {vat_amount}")
    for product in products:
        print(f"Product: {product.product.name}, Quantity: {product.quantity}, Unit Price: {product.unit_price}")

    context = {
        'coming': coming,
        'products': products,
        'total_amount': total_amount,
        'vat_amount': vat_amount,
    }

    # Рендеринг HTML в PDF
    html_string = render_to_string('warehouses/coming_pdf.html', context)
    pdf = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="coming_{coming_id}.pdf"'
    return response
